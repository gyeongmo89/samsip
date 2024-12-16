from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from . import models
from .database import engine, get_db, Base
from pydantic import BaseModel
from typing import Optional, List
import logging

from sqlalchemy.orm import joinedload
from datetime import datetime
import io
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create database tables
Base.metadata.create_all(bind=engine)

# Create FastAPI app
app = FastAPI()

# CORS 설정
origins = [
    "http://localhost:3000",
    "http://localhost:8000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create database tables


@app.on_event("startup")
async def startup_event():
    logger.info("Creating database tables...")
    try:
        Base.metadata.create_all(bind=engine)
        logger.info("Database tables created successfully!")
    except Exception as e:
        logger.error(f"Error during startup: {e}")
        raise e
    logger.info("Database initialization completed!")


# Pydantic models


class SupplierBase(BaseModel):
    name: str
    contact: Optional[str] = None
    address: Optional[str] = None


class SupplierResponse(SupplierBase):
    id: int

    class Config:
        from_attributes = True


class ItemBase(BaseModel):
    name: str
    description: Optional[str] = None
    price: Optional[float] = None


class ItemCreate(ItemBase):
    pass


class ItemResponse(ItemBase):
    id: int

    class Config:
        from_attributes = True


class UnitBase(BaseModel):
    name: str
    description: Optional[str] = None


class UnitCreate(UnitBase):
    pass


class UnitResponse(UnitBase):
    id: int

    class Config:
        from_attributes = True


class OrderBase(BaseModel):
    supplier_id: int
    item_id: int
    unit_id: int
    quantity: float
    price: float
    total: float
    payment_cycle: str
    payment_method: str
    client: str
    notes: Optional[str] = None
    date: Optional[str] = None


class OrderResponse(BaseModel):
    id: int
    date: Optional[str] = None
    supplier_id: int
    item_id: int
    unit_id: int
    quantity: float
    price: float
    total: float
    payment_cycle: str
    payment_method: str
    client: str
    notes: Optional[str] = None
    supplier: SupplierResponse
    item: ItemResponse
    unit: UnitResponse

    class Config:
        from_attributes = True


class OrderCreate(BaseModel):
    supplier_name: str
    item_name: str
    unit_name: str
    quantity: float
    price: float
    total: float
    payment_cycle: str
    payment_method: str
    client: str
    notes: Optional[str] = None
    date: Optional[str] = None


def get_float_value(value):
    if not value:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if value.startswith("="):  # 엑셀 수식인 경우
            return 0.0  # 기본값 반환 또는 다른 처리 로직 추가
        try:
            return float(value.replace(",", ""))  # 천단위 구분자 제거
        except ValueError:
            return 0.0
    return 0.0


# API endpoints


@app.post("/suppliers/", response_model=SupplierResponse)
def create_supplier(supplier: SupplierBase, db: Session = Depends(get_db)):
    try:
        # Check if supplier with same name already exists
        existing_supplier = (
            db.query(models.Supplier)
            .filter(
                models.Supplier.name == supplier.name,
                models.Supplier.is_deleted is False,
            )
            .first()
        )
        if existing_supplier:
            raise HTTPException(status_code=400, detail="already_exists")

        db_supplier = models.Supplier(
            name=supplier.name,
            contact=supplier.contact,
            address=supplier.address,
            is_deleted=False,
        )
        db.add(db_supplier)
        db.commit()
        db.refresh(db_supplier)
        return db_supplier
    except HTTPException as e:
        raise e
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating supplier: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/suppliers/", response_model=List[SupplierResponse])
def read_suppliers(db: Session = Depends(get_db)):
    suppliers = db.query(models.Supplier).all()
    return suppliers


@app.delete("/suppliers/bulk-delete")
def bulk_delete_suppliers(supplier_ids: List[int], db: Session = Depends(get_db)):
    try:
        # 실제로 데이터를 삭제
        db.query(models.Supplier).filter(models.Supplier.id.in_(supplier_ids)).delete(
            synchronize_session=False
        )
        db.commit()
        return {"message": "Suppliers deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/items/", response_model=ItemResponse)
def create_item(item: ItemCreate, db: Session = Depends(get_db)):
    try:
        # Check if item with same name already exists
        existing_item = (
            db.query(models.Item)
            .filter(models.Item.name == item.name, models.Item.is_deleted is False)
            .first()
        )
        if existing_item:
            raise HTTPException(status_code=400, detail="already_exists")

        # Convert price to float if it exists
        item_data = item.dict()
        if item_data.get("price") is not None:
            try:
                price = float(item_data["price"])
                if price < 0:
                    raise HTTPException(status_code=400, detail="price_invalid")
                item_data["price"] = price
            except ValueError:
                raise HTTPException(status_code=400, detail="price_invalid")

        db_item = models.Item(**item_data, is_deleted=False)
        db.add(db_item)
        db.commit()
        db.refresh(db_item)
        return db_item
    except HTTPException as e:
        raise e
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating item: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/items/", response_model=List[ItemResponse])
def read_items(db: Session = Depends(get_db)):
    items = db.query(models.Item).all()
    return items


@app.delete("/items/bulk-delete")
def bulk_delete_items(item_ids: List[int], db: Session = Depends(get_db)):
    try:
        db.query(models.Item).filter(models.Item.id.in_(item_ids)).delete(
            synchronize_session=False
        )
        db.commit()
        return {"message": "Items deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/units/", response_model=UnitResponse)
def create_unit(unit: UnitCreate, db: Session = Depends(get_db)):
    try:
        # Check if unit with same name already exists
        existing_unit = (
            db.query(models.Unit)
            .filter(models.Unit.name == unit.name, models.Unit.is_deleted is False)
            .first()
        )
        if existing_unit:
            raise HTTPException(status_code=400, detail="already_exists")

        # Convert None to empty string for description
        unit_data = unit.dict()
        if unit_data.get("description") is None:
            unit_data["description"] = ""

        db_unit = models.Unit(**unit_data)
        db.add(db_unit)
        db.commit()
        db.refresh(db_unit)
        return db_unit
    except HTTPException as e:
        raise e
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating unit: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/units/", response_model=List[UnitResponse])
def read_units(db: Session = Depends(get_db)):
    units = db.query(models.Unit).all()
    return units


@app.delete("/units/bulk-delete")
def bulk_delete_units(unit_ids: List[int], db: Session = Depends(get_db)):
    try:
        db.query(models.Unit).filter(models.Unit.id.in_(unit_ids)).delete(
            synchronize_session=False
        )
        db.commit()
        return {"message": "Units deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/orders/", response_model=List[OrderResponse])
def read_orders(db: Session = Depends(get_db)):
    try:
        orders = (
            db.query(models.Order)
            .filter(models.Order.is_deleted.is_(False))
            .options(
                joinedload(models.Order.supplier),
                joinedload(models.Order.item),
                joinedload(models.Order.unit),
            )
            .all()
        )
        order_responses = []

        for order in orders:
            # 공급처 정보
            supplier_data = {
                "id": order.supplier.id if order.supplier else -1,
                "name": order.supplier.name if order.supplier else "[삭제됨]",
                "contact": order.supplier.contact if order.supplier else None,
                "address": order.supplier.address if order.supplier else None,
            }

            # 품목 정보
            item_data = {
                "id": order.item.id if order.item else -1,
                "name": order.item.name if order.item else "[삭제됨]",
                "description": order.item.description if order.item else None,
                "price": order.item.price if order.item else None,
            }

            # 단위 정보
            unit_data = {
                "id": order.unit.id if order.unit else -1,
                "name": order.unit.name if order.unit else "[삭제됨]",
                "description": order.unit.description if order.unit else None,
            }

            # 주문 정보를 딕셔너리로 변환
            order_dict = {
                "id": order.id,
                "date": order.date,
                "supplier_id": order.supplier_id,
                "item_id": order.item_id,
                "unit_id": order.unit_id,
                "quantity": float(order.quantity),
                "price": float(order.price),  # 등록 당시의 가격 사용
                "total": float(order.total),  # 등록 당시의 총액 사용
                "payment_cycle": order.payment_cycle,
                "payment_method": order.payment_method,
                "client": order.client,
                "notes": order.notes,
                "supplier": supplier_data,
                "item": item_data,
                "unit": unit_data,
            }
            order_responses.append(order_dict)

        return order_responses
    except Exception as e:
        logger.error(f"Error retrieving orders: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/orders/", response_model=OrderResponse)
def create_order(order: OrderBase, db: Session = Depends(get_db)):
    try:
        db_order = models.Order(
            supplier_id=order.supplier_id,
            item_id=order.item_id,
            unit_id=order.unit_id,
            quantity=order.quantity,
            price=order.price,
            total=order.total,
            payment_cycle=order.payment_cycle,
            payment_method=order.payment_method,
            client=order.client,
            notes=order.notes,
            date=order.date,
        )
        db.add(db_order)
        db.commit()
        db.refresh(db_order)
        return db_order
    except Exception as e:
        logger.error(f"Error creating order: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/orders/{order_id}")
def update_order(order_id: int, order: OrderCreate, db: Session = Depends(get_db)):
    db_order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not db_order:
        raise HTTPException(status_code=404, detail="Order not found")

    # 구입처 찾기 또는 생성
    supplier = (
        db.query(models.Supplier)
        .filter(models.Supplier.name == order.supplier_name)
        .first()
    )
    if not supplier:
        supplier = models.Supplier(name=order.supplier_name)
        db.add(supplier)
        db.flush()

    # 품목 찾기 또는 생성
    item = db.query(models.Item).filter(models.Item.name == order.item_name).first()
    if not item:
        item = models.Item(name=order.item_name)
        db.add(item)
        db.flush()

    # 단위 찾기 또는 생성
    unit = db.query(models.Unit).filter(models.Unit.name == order.unit_name).first()
    if not unit:
        unit = models.Unit(name=order.unit_name)
        db.add(unit)
        db.flush()

    # 발주 데이터 업데이트
    for key, value in order.dict().items():
        if key not in ["supplier_name", "item_name", "unit_name"]:
            setattr(db_order, key, value)

    db_order.supplier_id = supplier.id
    db_order.item_id = item.id
    db_order.unit_id = unit.id

    db.commit()
    db.refresh(db_order)
    return db_order


@app.delete("/orders/bulk-delete")
def delete_orders(ids: List[int], db: Session = Depends(get_db)):
    try:
        for order_id in ids:
            order = db.query(models.Order).filter(models.Order.id == order_id).first()
            if order:
                db.delete(order)
        db.commit()
        return {"message": "Orders deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/orders/upload")
async def upload_orders(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    try:
        contents = await file.read()
        wb = load_workbook(
            filename=io.BytesIO(contents), data_only=True
        )  # data_only=True로 수식 대신 값을 가져옴
        ws = wb.active

        orders = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 빈 행 건너뛰기
                continue

            try:
                order_date = datetime.strptime(str(row[0]), "%Y-%m-%d").date()
            except ValueError:
                order_date = datetime.now().date()
                order_date = datetime.now().date()

            # 구입처 찾기 또는 생성
            supplier = (
                db.query(models.Supplier).filter(models.Supplier.name == row[1]).first()
            )
            if not supplier:
                supplier = models.Supplier(name=row[1], contact=row[9])
                db.add(supplier)
                db.flush()

            # 품목 찾기 또는 생성
            item = db.query(models.Item).filter(models.Item.name == row[2]).first()
            if not item:
                item = models.Item(name=row[2], price=get_float_value(row[3]))
                db.add(item)
                db.flush()

            # 단위 찾기 또는 생성
            unit = db.query(models.Unit).filter(models.Unit.name == row[4]).first()
            if not unit:
                unit = models.Unit(name=row[4])
                db.add(unit)
                db.flush()

            # 발주 데이터 생성
            order = models.Order(
                date=order_date,
                supplier_id=supplier.id,
                item_id=item.id,
                unit_id=unit.id,
                price=get_float_value(row[3]),
                quantity=get_float_value(row[5]),
                total=get_float_value(row[6]),
                payment_cycle=str(row[7] or ""),
                payment_method=str(row[8] or ""),
                client=str(row[9] or ""),
                notes=str(row[10] or ""),
            )
            orders.append(order)

        db.add_all(orders)
        db.commit()

        return {"message": f"Successfully uploaded {len(orders)} orders"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
def update_supplier(
    supplier_id: int, supplier: SupplierBase, db: Session = Depends(get_db)
):
    db_supplier = (
        db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    )
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    for key, value in supplier.dict().items():
        setattr(db_supplier, key, value)

    db.commit()
    db.refresh(db_supplier)
    return db_supplier


@app.put("/items/{item_id}", response_model=ItemResponse)
def update_item(item_id: int, item: ItemBase, db: Session = Depends(get_db)):
    db_item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")

    for key, value in item.dict().items():
        setattr(db_item, key, value)

    db.commit()
    db.refresh(db_item)
    return db_item


@app.put("/units/{unit_id}", response_model=UnitResponse)
def update_unit(unit_id: int, unit: UnitCreate, db: Session = Depends(get_db)):
    try:
        db_unit = db.query(models.Unit).filter(models.Unit.id == unit_id).first()
        if not db_unit:
            raise HTTPException(status_code=404, detail="Unit not found")

        # Check if another unit with the same name exists
        existing_unit = (
            db.query(models.Unit)
            .filter(
                models.Unit.name == unit.name,
                models.Unit.id != unit_id,
                models.Unit.is_deleted is False,
            )
            .first()
        )
        if existing_unit:
            raise HTTPException(status_code=400, detail="already_exists")

        # Convert None to empty string for description
        unit_data = unit.dict()
        if unit_data.get("description") is None:
            unit_data["description"] = ""

        for key, value in unit_data.items():
            setattr(db_unit, key, value)

        db.commit()
        db.refresh(db_unit)
        return db_unit
    except HTTPException as e:
        raise e
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating unit: {e}")
        raise HTTPException(status_code=500, detail=str(e))
